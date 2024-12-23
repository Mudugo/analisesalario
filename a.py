import openpyxl

caminho_arquivo = r'Desktop\incon\123.xlsx'
wb = openpyxl.load_workbook(caminho_arquivo)


ws_origem = wb.active
planilhas_cargos = {}

for row in ws_origem.iter_rows(min_row=2, values_only=True):
    cargo = row[4]

    if cargo:
        if cargo not in planilhas_cargos:
            ws_nova = wb.create_sheet(title=cargo[:31])
            planilhas_cargos[cargo] = ws_nova

            for i, cell in enumerate(ws_origem[1], start=1):
                ws_nova.cell(row=1, column=i, value=cell.value)

        ws_cargo = planilhas_cargos[cargo]
        ws_cargo.append(row)

wb.save('Analise de Cargos.xlsx')
